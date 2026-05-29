#!/usr/bin/env python3
"""Группировка типов организаций из final/final.md в 10 категорий → Excel."""

from __future__ import annotations

from collections import Counter, defaultdict
from pathlib import Path

from org_type_groups import _RULES, classify

ROOT = Path(__file__).resolve().parents[1]
FINAL_MD = ROOT / "final" / "final.md"
OUT_XLSX = ROOT / "final" / "final.xlsx"

MAIN_TYPES = [name for name, _ in _RULES]


def load_type_counts() -> Counter[str]:
    counts: Counter[str] = Counter()
    for ln in FINAL_MD.read_text(encoding="utf-8").splitlines():
        if not ln.startswith("|") or "Организация" in ln or ln.startswith("|---"):
            continue
        cells = [c.strip() for c in ln.strip("|").split("|")]
        if len(cells) >= 2:
            counts[cells[1]] += 1
    return counts


def main() -> int:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    type_counts = load_type_counts()
    mapping: dict[str, str] = {t: classify(t) for t in type_counts}

    group_counts: Counter[str] = Counter()
    group_types: dict[str, list[tuple[str, int]]] = defaultdict(list)
    for t, cnt in type_counts.items():
        g = mapping[t]
        group_counts[g] += cnt
        group_types[g].append((t, cnt))

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "10 основных типов"
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")
    headers = ["№", "Основной тип", "Записей", "Доля %", "Подтипов"]
    ws1.append(headers)
    for cell in ws1[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    total = sum(type_counts.values())
    for i, g in enumerate(MAIN_TYPES, 1):
        cnt = group_counts[g]
        pct = round(100 * cnt / total, 1) if total else 0
        ws1.append([i, g, cnt, pct, len(group_types[g])])

    ws1.append([])
    ws1.append(["", "ИТОГО", total, 100, len(type_counts)])
    ws1.column_dimensions["A"].width = 5
    ws1.column_dimensions["B"].width = 32
    ws1.column_dimensions["C"].width = 12
    ws1.column_dimensions["D"].width = 10
    ws1.column_dimensions["E"].width = 12

    ws2 = wb.create_sheet("Все подтипы")
    h2 = ["Основной тип", "Подтип (как в final.md)", "Записей"]
    ws2.append(h2)
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font
    rows = []
    for g in MAIN_TYPES:
        for t, cnt in sorted(group_types[g], key=lambda x: (-x[1], x[0].casefold())):
            rows.append([g, t, cnt])
    for r in rows:
        ws2.append(r)
    ws2.column_dimensions["A"].width = 32
    ws2.column_dimensions["B"].width = 48
    ws2.column_dimensions["C"].width = 12
    ws2.auto_filter.ref = f"A1:C{len(rows)+1}"

    ws3 = wb.create_sheet("Группы детально")
    ws3.append(["Основной тип", "Подтипы"])
    for cell in ws3[1]:
        cell.fill = header_fill
        cell.font = header_font
    for g in MAIN_TYPES:
        parts = sorted(group_types[g], key=lambda x: (-x[1], x[0].casefold()))
        lines = [f"{t} ({c})" for t, c in parts]
        ws3.append([g, "\n".join(lines)])
    ws3.column_dimensions["A"].width = 32
    ws3.column_dimensions["B"].width = 80
    for row in ws3.iter_rows(min_row=2, max_col=2):
        row[1].alignment = Alignment(wrap_text=True, vertical="top")

    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)
    print(f"Wrote {OUT_XLSX}")
    print(f"Total rows in final.md: {total}")
    print(f"Unique subtypes: {len(type_counts)}")
    for g in MAIN_TYPES:
        print(f"  {g}: {group_counts[g]} ({len(group_types[g])} subtypes)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
